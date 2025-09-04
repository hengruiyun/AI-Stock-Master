"""
from config.gui_i18n import t_gui as _
# t_common 功能由 t_gui 替代
t_common = t_gui
报告生成器模块 - 修复版本

提供Excel、HTML、PDF等格式的分析报告生成功能。

主要功能:
- 生成完整的分析报告
- 支持多种格式输出
- 集成图表和数据表
- 自动样式美化

作者: 267278466@qq.com
版本: 1.0.0
"""

import os
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union
import json
import logging

# 添加项目根目录到Python路径以导入配置
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
sys.path.insert(0, project_root)

# 导入国际化翻译
from config.gui_i18n import t_gui

# 数据处理
import pandas as pd
import numpy as np

# Excel处理
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing import image
from openpyxl.chart import LineChart, Reference

# 配置日志
logger = logging.getLogger(__name__)

# 图表相关导入 (可选)
try:
    import plotly.graph_objects as go
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False
    logger.warning(t_gui('plotly_not_installed'))

class ReportGenerator:
    """主报告生成器"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # 报告配置
        self.config = {
            'title': 'AI股票分析报告',
            'subtitle': '智能量化分析系统',
            'theme': 'professional',
            'include_charts': True,
            'include_data_tables': True,
            'include_recommendations': True,
            'chart_formats': ['html', 'png'],
            'logo_path': None
        }
    
    def _get_html_lang(self):
        """获取HTML语言标识"""
        try:
            from config.gui_i18n import get_system_language
            return "en" if get_system_language() == 'en' else "zh-CN"
        except:
            return "zh-CN"
    
    def generate_complete_report(self, analysis_data: Dict[str, Any], 
                               formats: List[str] = ['html', 'excel']) -> Dict[str, str]:
        """
        生成完整的分析报告
        
        Args:
            analysis_data: 分析数据字典
            formats: 报告格式列表 ['html', 'excel', 'pdf']
            
        Returns:
            生成的报告文件路径字典
        """
        
        logger.info(t_gui('starting_complete_report'))
        
        # 准备报告数据
        report_data = self._prepare_report_data(analysis_data)
        
        # 生成各种格式的报告
        generated_files = {}
        
        if 'html' in formats:
            html_path = self._generate_html_report(report_data)
            generated_files['html'] = html_path
        
        if 'excel' in formats:
            excel_path = self._generate_excel_report(report_data)
            generated_files['excel'] = excel_path
        
        logger.info(t_gui('report_generation_complete', count=len(generated_files)))
        return generated_files
    
    def _format_stock_code_for_display(self, stock_code: str) -> str:
        """格式化股票代码用于显示
        
        Args:
            stock_code: 原始股票代码
            
        Returns:
            格式化后的股票代码
        """
        if not stock_code:
            return stock_code
            
        # 判断是否为中国市场股票代码（6位纯数字）
        if stock_code.isdigit() and len(stock_code) == 6:
            # 中国市场保持完整6位数
            return stock_code
        else:
            # 其他市场去除前导0
            return stock_code.lstrip('0') if stock_code.lstrip('0') else '0'
    
    def _prepare_report_data(self, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """准备报告数据"""
        
        report_data = {
            'metadata': {
                'generated_at': datetime.now(),
                'data_date': analysis_data.get('data_date', datetime.now().date()),
                'total_stocks': len(analysis_data.get('stocks', {})),
                'total_industries': len(analysis_data.get('industries', {})),
                'analysis_period': '实际分析周期',
                'system_version': '1.0.0'
            },
            'stocks': analysis_data.get('stocks', {}),
            'industries': analysis_data.get('industries', {}),
            'market': analysis_data.get('market', {}),
            'performance': analysis_data.get('performance', {}),
            'summary': self._generate_summary_stats(analysis_data)
        }
        
        return report_data
    
    def _generate_summary_stats(self, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """生成汇总统计 - 只使用真实数据"""
        
        stocks = analysis_data.get('stocks', {})
        
        # 只使用真实RTSI数据，不生成模拟数据
        if not stocks:
            return {
                'market_overview': {'total_analyzed': 0},
                'market_sentiment': {'sentiment_score': 0},
                'top_performers': [],
                'risk_analysis': {'market_stability': 'no_data'}
            }
        
        rtsi_values = []
        for stock_info in stocks.values():
            # 从真实分析结果获取RTSI值
            rtsi_data = stock_info.get('rtsi', {})
            if isinstance(rtsi_data, dict):
                rtsi = rtsi_data.get('rtsi', 0)
            else:
                rtsi = rtsi_data if isinstance(rtsi_data, (int, float)) else 0
            
            if rtsi > 0:  # 只包含有效的RTSI值
                rtsi_values.append(rtsi)
        
        if rtsi_values:
            summary = {
                'market_overview': {
                    'total_analyzed': len(stocks),
                    'valid_rtsi_count': len(rtsi_values),
                    'average_rtsi': np.mean(rtsi_values),
                    'max_rtsi': np.max(rtsi_values),
                    'min_rtsi': np.min(rtsi_values),
                    'rtsi_std': np.std(rtsi_values)
                },
                'market_sentiment': {
                    'bullish_stocks': sum(1 for r in rtsi_values if r > 70),
                    'neutral_stocks': sum(1 for r in rtsi_values if 30 <= r <= 70),
                    'bearish_stocks': sum(1 for r in rtsi_values if r < 30),
                    'sentiment_score': np.mean(rtsi_values)
                },
                'top_performers': self._get_top_performers(stocks, rtsi_values),
                'risk_analysis': {
                    'high_risk_stocks': sum(1 for r in rtsi_values if r > 80 or r < 20),
                    'volatility_index': np.std(rtsi_values),
                    'market_stability': 'stable' if np.std(rtsi_values) < 25 else 'volatile'
                }
            }
        else:
            summary = {
                'market_overview': {'total_analyzed': len(stocks), 'valid_rtsi_count': 0},
                'market_sentiment': {'sentiment_score': 0},
                'top_performers': [],
                'risk_analysis': {'market_stability': 'insufficient_data'}
            }
        
        return summary
    
    def _get_top_performers(self, stocks: Dict[str, Any], 
                          rtsi_values: List[float]) -> List[Dict[str, Any]]:
        """获取表现最佳的股票"""
        
        if not stocks or not rtsi_values:
            return []
        
        # 从真实数据创建排名
        stock_rtsi_pairs = []
        for code, stock_info in stocks.items():
            rtsi_data = stock_info.get('rtsi', {})
            if isinstance(rtsi_data, dict):
                rtsi = rtsi_data.get('rtsi', 0)
            else:
                rtsi = rtsi_data if isinstance(rtsi_data, (int, float)) else 0
            
            if rtsi > 0:  # 只包含有效的RTSI值
                stock_rtsi_pairs.append((code, rtsi))
        
        # 按RTSI排序，取前10
        top_performers = sorted(stock_rtsi_pairs, key=lambda x: x[1], reverse=True)[:10]
        
        result = []
        for code, rtsi in top_performers:
            stock_info = stocks.get(code, {})
            result.append({
                'code': code,
                'name': stock_info.get('name', ''),
                'industry': stock_info.get('industry', '未分类'),
                'rtsi': rtsi,
                'trend': 'strong_up' if rtsi > 80 else 'up' if rtsi > 60 else 'neutral'
            })
        
        return result
    
    def _generate_html_report(self, report_data: Dict[str, Any]) -> str:
        """生成HTML报告"""
        
        try:
            output_path = self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            
            html_content = f"""
<!DOCTYPE html>
<html lang="{self._get_html_lang()}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.config['title']}</title>
    <style>
        body {{ font-family: 'Microsoft YaHei', sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 30px; }}
        .title {{ color: #2c3e50; font-size: 28px; margin-bottom: 10px; }}
        .subtitle {{ color: #7f8c8d; font-size: 16px; }}
        .section {{ margin-bottom: 30px; }}
        .section-title {{ color: #34495e; font-size: 20px; border-bottom: 2px solid #3498db; padding-bottom: 8px; margin-bottom: 15px; }}
        .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
        .metric-card {{ background: #f8f9fa; padding: 15px; border-radius: 6px; text-align: center; }}
        .metric-value {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
        .metric-label {{ font-size: 14px; color: #7f8c8d; margin-top: 5px; }}
        .table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
        .table th, .table td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
        .table th {{ background-color: #f8f9fa; font-weight: bold; }}
        .footer {{ text-align: center; margin-top: 30px; color: #7f8c8d; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1 class="title">{self.config['title']}</h1>
            <p class="subtitle">{self.config['subtitle']}</p>
            <p>{t_gui('generation_time')} {report_data['metadata']['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="section">
            <h2 class="section-title">{t_gui('data_overview')}</h2>
            <div class="metric-grid">
                <div class="metric-card">
                    <div class="metric-value">{report_data['metadata']['total_stocks']}</div>
                    <div class="metric-label">总股票数</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{report_data['metadata']['total_industries']}</div>
                    <div class="metric-label">行业数量</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{report_data['summary']['market_overview'].get('valid_rtsi_count', 0)}</div>
                    <div class="metric-label">有效分析</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{report_data['summary']['market_overview'].get('average_rtsi', 0):.1f}</div>
                    <div class="metric-label">平均RTSI</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2 class="section-title">市场情绪</h2>
            <div class="metric-grid">
                <div class="metric-card">
                    <div class="metric-value" style="color: #27ae60;">{report_data['summary']['market_sentiment'].get('bullish_stocks', 0)}</div>
                    <div class="metric-label">看多股票</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value" style="color: #f39c12;">{report_data['summary']['market_sentiment'].get('neutral_stocks', 0)}</div>
                    <div class="metric-label">中性股票</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value" style="color: #e74c3c;">{report_data['summary']['market_sentiment'].get('bearish_stocks', 0)}</div>
                    <div class="metric-label">看空股票</div>
                </div>
            </div>
        </div>
        
        <div class="section">
            <h2 class="section-title">优质股票排行</h2>
            <table class="table">
                <thead>
                    <tr>
                        <th>排名</th>
                        <th>股票代码</th>
                        <th>股票名称</th>
                        <th>所属行业</th>
                        <th>RTSI指数</th>
                        <th>趋势状态</th>
                    </tr>
                </thead>
                <tbody>
"""
            
            # 添加股票排行数据
            for i, stock in enumerate(report_data['summary']['top_performers'][:10], 1):
                # 格式化股票代码：中国市场保持6位数，其他市场去除前导0
                stock_code = stock.get('code', '')
                formatted_code = self._format_stock_code_for_display(stock_code)
                
                html_content += f"""
                    <tr>
                        <td>{i}</td>
                        <td>{formatted_code}</td>
                        <td>{stock.get('name', '')}</td>
                        <td>{stock.get('industry', '')}</td>
                        <td>{stock.get('rtsi', 0):.2f}</td>
                        <td>{stock.get('trend', '')}</td>
                    </tr>
"""
            
            html_content += f"""
                </tbody>
            </table>
        </div>
        
        <div class="section">
            <h2 class="section-title">行业轮动分析</h2>
            <p style="color: #7f8c8d; margin-bottom: 15px;">基于TMA的行业轮动分析</p>
            <table class="table">
                <thead>
                    <tr>
                        <th>排名</th>
                        <th>行业名称</th>
                        <th>TMA指数</th>
                        <th>强度等级</th>
                        <th>投资建议</th>
                    </tr>
                </thead>
                <tbody>
"""
            
            # 添加行业轮动分析数据
            if 'industries' in report_data and report_data['industries']:
                industries_list = []
                for industry_name, industry_info in report_data['industries'].items():
                    # 提取TMA值（存储在irsi键中）
                    irsi_value = industry_info.get('irsi', 0)
                    if isinstance(irsi_value, dict):
                        tma_value = irsi_value.get('irsi', 0)
                    else:
                        tma_value = irsi_value
                    
                    if isinstance(tma_value, (int, float)) and tma_value != 0:
                        # 判断强度等级和投资建议
                        if tma_value > 20:
                            strength = "强势"
                            advice = "积极配置"
                            color = "#27ae60"
                        elif tma_value > 5:
                            strength = "中性偏强"
                            advice = "适度关注"
                            color = "#f39c12"
                        elif tma_value > -5:
                            strength = "中性"
                            advice = "观望"
                            color = "#95a5a6"
                        elif tma_value > -20:
                            strength = "中性偏弱"
                            advice = "谨慎"
                            color = "#e67e22"
                        else:
                            strength = "弱势"
                            advice = "回避"
                            color = "#e74c3c"
                        
                        industries_list.append({
                            'name': industry_name,
                            'tma': tma_value,
                            'strength': strength,
                            'advice': advice,
                            'color': color
                        })
                
                # 按TMA值排序
                industries_list.sort(key=lambda x: x['tma'], reverse=True)
                
                # 显示前5名行业
                for i, industry in enumerate(industries_list[:5], 1):
                    html_content += f"""
                    <tr>
                        <td>{i}</td>
                        <td>{industry['name']}</td>
                        <td style="color: {industry['color']}; font-weight: bold;">{industry['tma']:.2f}</td>
                        <td style="color: {industry['color']};">{industry['strength']}</td>
                        <td>{industry['advice']}</td>
                    </tr>
"""
                
                if industries_list:
                    strongest_industry = industries_list[0]
                    html_content += f"""
                </tbody>
            </table>
            <p style="margin-top: 15px; color: #34495e;">
                <strong>当前最强行业:</strong> {strongest_industry['name']} (TMA指数: {strongest_industry['tma']:.2f})<br/>
                <small style="color: #7f8c8d;">TMA指数说明：正值表示行业相对强势，负值表示相对弱势</small>
            </p>
"""
                else:
                    html_content += f"""
                    <tr>
                        <td colspan="5" style="text-align: center; color: #7f8c8d; padding: 20px;">暂无行业分析数据</td>
                    </tr>
                </tbody>
            </table>
"""
            else:
                html_content += f"""
                    <tr>
                        <td colspan="5" style="text-align: center; color: #7f8c8d; padding: 20px;">暂无行业分析数据</td>
                    </tr>
                </tbody>
            </table>
"""
            
            html_content += """
        </div>
            
        
        <div class="footer">
            <p>{t_common('report_auto_generated')} | {t_common('version')}: {report_data['metadata']['system_version']}</p>
            <p>{t_common('analysis_disclaimer')}</p>
        </div>
    </div>
</body>
</html>
"""
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"HTML报告生成成功: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"生成HTML报告失败: {e}")
            return ""
    
    def _generate_excel_report(self, report_data: Dict[str, Any]) -> str:
        """生成Excel报告"""
        
        try:
            output_path = self.output_dir / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = ""
            
            # 标题
            ws['A1'] = self.config['title']
            ws['A1'].font = Font(name='Microsoft YaHei', size=18, bold=True, color='2E86AB')
            
            # 基本信息
            row = 3
            ws[f'A{row}'] = "生成时间:"
            ws[f'B{row}'] = report_data['metadata']['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
            row += 1
            
            ws[f'A{row}'] = "总股票数:"
            ws[f'B{row}'] = report_data['metadata']['total_stocks']
            row += 1
            
            ws[f'A{row}'] = "行业数量:"
            ws[f'B{row}'] = report_data['metadata']['total_industries']
            row += 2
            
            # 优质股票排行
            ws[f'A{row}'] = "优质股票排行 (按RTSI排序)"
            ws[f'A{row}'].font = Font(name='Microsoft YaHei', size=14, bold=True)
            row += 1
            
            headers = ['排名', '股票代码', '股票名称', '所属行业', 'RTSI指数', '趋势状态']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
                ws.cell(row=row, column=col).font = Font(name='Microsoft YaHei', bold=True)
            row += 1
            
            # 修复：正确处理top_performers数据
            top_performers = report_data['summary'].get('top_performers', [])
            if not top_performers and 'stocks' in report_data:
                # 如果没有top_performers，从stocks数据生成
                top_performers = []
                for code, stock_info in report_data['stocks'].items():
                    rtsi_value = stock_info.get('rtsi', 0)
                    if isinstance(rtsi_value, dict):
                        rtsi_value = rtsi_value.get('rtsi', 0)
                    
                    top_performers.append({
                        'code': code,
                        'name': stock_info.get('name', code),
                        'industry': stock_info.get('industry', '未分类'),
                        'rtsi': rtsi_value,
                        'trend': '未知'
                    })
                
                # 按RTSI排序
                top_performers.sort(key=lambda x: x.get('rtsi', 0), reverse=True)
            
            # 写入数据行
            for i, stock in enumerate(top_performers[:20], 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=stock.get('code', ''))
                ws.cell(row=row, column=3, value=stock.get('name', ''))
                ws.cell(row=row, column=4, value=stock.get('industry', ''))
                ws.cell(row=row, column=5, value=float(stock.get('rtsi', 0)))
                ws.cell(row=row, column=6, value=stock.get('trend', ''))
                row += 1
            
            # 添加行业分析工作表
            row += 2
            ws[f'A{row}'] = "强势行业排行 (按IRSI排序)"
            ws[f'A{row}'].font = Font(name='Microsoft YaHei', size=14, bold=True)
            row += 1
            
            industry_headers = ['排名', '行业名称', 'IRSI指数', '强度等级', '股票数量']
            for col, header in enumerate(industry_headers, 1):
                ws.cell(row=row, column=col, value=header)
                ws.cell(row=row, column=col).font = Font(name='Microsoft YaHei', bold=True)
            row += 1
            
            # 处理行业数据
            industries_data = []
            if 'industries' in report_data:
                for industry_name, industry_info in report_data['industries'].items():
                    irsi_value = industry_info.get('irsi', 0)
                    if isinstance(irsi_value, dict):
                        irsi_value = irsi_value.get('irsi', 0)
                    
                    # 判断强度等级
                    if irsi_value > 20:
                        strength = "强势"
                    elif irsi_value > 5:
                        strength = "中性偏强"
                    elif irsi_value > -5:
                        strength = "中性"
                    elif irsi_value > -20:
                        strength = "中性偏弱"
                    else:
                        strength = "弱势"
                    
                    industries_data.append({
                        'name': industry_name,
                        'irsi': irsi_value,
                        'strength': strength,
                        'stock_count': industry_info.get('stock_count', 0)
                    })
                
                # 按IRSI排序
                industries_data.sort(key=lambda x: x['irsi'], reverse=True)
                
                # 写入行业数据
                for i, industry in enumerate(industries_data[:15], 1):
                    ws.cell(row=row, column=1, value=i)
                    ws.cell(row=row, column=2, value=industry['name'])
                    ws.cell(row=row, column=3, value=float(industry['irsi']))
                    ws.cell(row=row, column=4, value=industry['strength'])
                    ws.cell(row=row, column=5, value=industry['stock_count'])
                    row += 1
            
            wb.save(output_path)
            logger.info(f"Excel报告生成成功: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")
            return ""

class ExcelReportGenerator:
    """Excel格式报告生成器"""
    
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
    
    def create_report(self, report_data: Dict[str, Any]) -> str:
        """创建Excel报告"""
        try:
            output_path = self.output_dir / f"excel_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = ""
            
            # 报告标题和时间
            ws['A1'] = "AI股票大师分析报告"
            ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 基本统计信息
            metadata = report_data.get('metadata', {})
            ws['A4'] = f"总股票数: {metadata.get('total_stocks', 0)}"
            ws['A5'] = f"行业数量: {metadata.get('total_industries', 0)}"
            ws['A6'] = f"分析周期: {metadata.get('analysis_period', '未知')}"
            
            # 添加股票数据
            if 'stocks' in report_data and report_data['stocks']:
                ws['A8'] = "优质股票推荐 (按RTSI排序)"
                
                # 表头
                headers = ['排名', '股票代码', '股票名称', '所属行业', 'RTSI指数']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=9, column=col, value=header)
                
                # 数据行
                stocks_list = []
                for code, stock_info in report_data['stocks'].items():
                    rtsi_value = stock_info.get('rtsi', 0)
                    if isinstance(rtsi_value, dict):
                        rtsi_value = rtsi_value.get('rtsi', 0)
                    
                    stocks_list.append({
                        'code': code,
                        'name': stock_info.get('name', code),
                        'industry': stock_info.get('industry', '未分类'),
                        'rtsi': float(rtsi_value) if rtsi_value else 0.0
                    })
                
                # 按RTSI排序
                stocks_list.sort(key=lambda x: x['rtsi'], reverse=True)
                
                # 写入前20只股票
                for i, stock in enumerate(stocks_list[:20], 1):
                    row = 9 + i
                    ws.cell(row=row, column=1, value=i)
                    ws.cell(row=row, column=2, value=stock['code'])
                    ws.cell(row=row, column=3, value=stock['name'])
                    ws.cell(row=row, column=4, value=stock['industry'])
                    ws.cell(row=row, column=5, value=stock['rtsi'])
            
            # 添加行业数据
            if 'industries' in report_data and report_data['industries']:
                start_row = 32
                ws[f'A{start_row}'] = "强势行业分析 (按IRSI排序)"
                
                # 表头
                industry_headers = ['排名', '行业名称', 'IRSI指数', '强度等级', '股票数量']
                for col, header in enumerate(industry_headers, 1):
                    ws.cell(row=start_row+1, column=col, value=header)
                
                # 数据处理
                industries_list = []
                for industry_name, industry_info in report_data['industries'].items():
                    irsi_value = industry_info.get('irsi', 0)
                    if isinstance(irsi_value, dict):
                        irsi_value = irsi_value.get('irsi', 0)
                    
                    # 判断强度等级
                    if irsi_value > 20:
                        strength = "强势"
                    elif irsi_value > 5:
                        strength = "中性偏强"
                    elif irsi_value > -5:
                        strength = "中性"
                    elif irsi_value > -20:
                        strength = "中性偏弱"
                    else:
                        strength = "弱势"
                    
                    industries_list.append({
                        'name': industry_name,
                        'irsi': float(irsi_value) if irsi_value else 0.0,
                        'strength': strength,
                        'stock_count': industry_info.get('stock_count', 0)
                    })
                
                # 按IRSI排序
                industries_list.sort(key=lambda x: x['irsi'], reverse=True)
                
                # 写入前15个行业
                for i, industry in enumerate(industries_list[:15], 1):
                    row = start_row + 1 + i
                    ws.cell(row=row, column=1, value=i)
                    ws.cell(row=row, column=2, value=industry['name'])
                    ws.cell(row=row, column=3, value=industry['irsi'])
                    ws.cell(row=row, column=4, value=industry['strength'])
                    ws.cell(row=row, column=5, value=industry['stock_count'])
            
            wb.save(output_path)
            logger.info(f"Excel报告生成成功: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Excel报告生成失败: {e}")
            return ""

# 便捷函数
def generate_report_quick(analysis_data: Dict[str, Any], 
                         formats: List[str] = ['html']) -> Dict[str, str]:
    """快速生成报告"""
    generator = ReportGenerator()
    return generator.generate_complete_report(analysis_data, formats)

def create_excel_report_quick(data: Dict[str, Any]) -> str:
    """快速生成Excel报告"""
    generator = ExcelReportGenerator(Path("reports"))
    return generator.create_report(data)

if __name__ == "__main__":
    # 测试代码
    print("-")
    
    # 创建测试数据
    test_data = {
        'stocks': {
            '600036': {'name': '招商银行', 'industry': '银行', 'rtsi': {'rtsi': 85.2}},
            '000001': {'name': '平安银行', 'industry': '银行', 'rtsi': {'rtsi': 72.1}}
        },
        'industries': {},
        'market': {},
        'data_date': datetime.now().date()
    }
    
    # 生成测试报告
    generator = ReportGenerator("test_reports")
    result = generator.generate_complete_report(test_data, ['html', 'excel'])
    
    print(f"测试报告生成完成: {result}")